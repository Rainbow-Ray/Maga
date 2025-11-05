import calendar
import openpyxl as ex
import openpyxl.styles
import pandas as pd
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image


class CalendarConfig:
    """Конфигурация календаря"""
    year: int = 2025
    month: int = 11

    start_column: int = 2
    start_row: int = 2
    end_column: int = start_column + 7
    end_row: int = start_row + 6

    day_row: int = start_row+1
    body_row: int = start_row+2


    left_edge_column = start_column
    right_edge_column = start_column + 7



    # Цвета
    black: str = 'FF000000'
    light_gray: str = 'D8D8D8'
    gray: str = 'A5A5A5'

    # Шрифты
    main_font: openpyxl.styles.Font = Font(name='Century Gothic',
                size=20,
                bold=False,
                italic=False,
                vertAlign='baseline',
                underline='none',
                strike=False,
                color='FF000000')

    odd_font = Font(name='Century Gothic',
                   size=20,
                   bold=False,
                   italic=False,
                   vertAlign='baseline',
                   underline='none',
                   strike=False,
                   color=gray)

    center_align = Alignment(horizontal='center',
                            vertical='center',
                            text_rotation=0,
                            wrap_text=False,
                            shrink_to_fit=False,
                            indent=0)

    # Размеры
    first_column_width: float = 25.6
    calendar_column_width: float = 13.6
    row_height: float = 59.4


class WriterConfig(CalendarConfig):
    start_column: int = 2
    start_row: int = 2

    day_row: int = start_row+1
    body_row: int = start_row+2

    left_edge_column = start_column
    right_edge_column = start_column + 7

    first_column_width: float = 25.6
    calendar_column_width: float = 13.6
    row_height: float = 59.4

    wb: ex.Workbook() = ex.Workbook()
    ws: wb.active = wb.active

    page_margins: PageMargins(left=0, right=0, top=0, bottom=0, header=0, footer=0)
    page_orientation:str = ws.ORIENTATION_LANDSCAPE
    paper_Size: ws.PAPERSIZE_A4

    name: str = "pandas_openpyxl.xlsx"






class _Calendar:
    DAYS_COLUMNS = {
        0: 'Пн', 1: 'Вт', 2: 'Ср', 3: 'Чт',
        4: 'Пт', 5: 'Сб', 6: 'Вс'}

    header: str = ''
    df: pd.DataFrame
    odd_month_indexes = {'Old': [-1, -1], 'New': [-1, -1]}
    odd_old_start = -1
    odd_old_end = -1
    odd_new_start = -1
    odd_new_end = -1

    height = 0
    width = 0

    def __init__(self, config: CalendarConfig):
        self.config = config
        self.df = self.calendar_df()
        self.header = getRuMonth(self.config.year, self.config.month)
        self.odd_month_indexes = self.get_odd_month_indexes()

        self.odd_old_start = self.odd_month_indexes['Old'][0] + self.config.start_column
        self.odd_old_end = self.odd_month_indexes['Old'][1] + self.config.start_column
        self.odd_new_start = self.odd_month_indexes['New'][0] + self.config.start_column
        self.odd_new_end = self.odd_month_indexes['New'][1] + self.config.start_column

        self.height = len(self.df)
        self.width = len(self.df.columns)
        config.end_column =  config.start_column + self.width
        config.end_row =  config.start_row + self.height

        self.styler = CalendarStyler(config)
        self.borders = self.styler.create_calendar_borders()

    def has_odd_at_start(self):
        if self.odd_month_indexes['Old'][0] > 0:
            return True
        return False

    def has_odd_at_end(self):
        if self.odd_month_indexes['New'][0] > 0:
            return True
        return False

    def format_to_day(self):
        df = pd.DataFrame()
        for h in self.df.columns:
            df[h] = self.df.apply(lambda cell: cell[h].day, axis=1)
        return df

    def get_odd_month_indexes(self):
        oldS, oldF = self.odd_month_in_row(0)
        newS, newF = self.odd_month_in_row(-1)
        return {'Old': [oldS, oldF], 'New': (newS, newF)}

    def odd_month_in_row(self, row: int):
        s = -1
        f = -1
        for i in range(self.width):
            c = self.df.iloc[row, i]
            if not isNowMonth(c.month):
                if s == -1:
                    s = i
                f = i
        return s, f

    def calendar_df(self):
        cal = calendar.Calendar()
        month = cal.monthdatescalendar(self.config.year, self.config.month)
        df = pd.DataFrame(month)
        df = df.rename(columns=self.DAYS_COLUMNS)
        return df

    def generateCalendar(self):
        df = self.format_to_day()

        designCal(ws, oddIndex)

        img = Image('4.jpg')
        ws.add_image(img, 'A1')

        wb.save("pandas_openpyxl.xlsx")


class borderChunk:
    start_col: int
    start_row: int
    width: int
    height: int

    def __init__(self, col, row, w, h):
        self.start_col =col
        self.start_row =row
        self.width =w
        self.height =h



class CalendarStyler:
    def __init__(self, config: CalendarConfig, cal: _Calendar, ws: ex.worksheet, wb: ex.Workbook):
        self.config = config
        self.calendar = cal
        self.ws = ws
        self.wb = wb

    def designCal(self):
        self.ws.merge_cells(f'{self.config.start_column}{self.config.start_row}:{self.config.end_column}{self.config.start_row}')
        self.set_dimensions()
        self.design_odd()

        self.designBorder(ws)

    def design_odd(self):
        if self.calendar.has_odd_at_start():
            self.apply_font(self.calendar.odd_old_start, self.calendar.odd_old_end, self.config.body_row,self.config.body_row,  self.config.odd_font)
        if self.calendar.has_odd_at_end():
            self.apply_font(self.calendar.odd_new_start, self.calendar.odd_new_end, self.config.end_row, self.config.end_row,  self.config.odd_font)

    def apply_font(self, start_col, finish_col, start_row, finish_row, style):
        for i in range(start_row, finish_row + 1):
            for j in range(start_col, finish_col +1):
                self.ws.cell(column=j, row=i).font = style

    def set_dimensions(self):
        self.ws.column_dimensions['A'].width = self.config.first_column_width
        for c in range(self.config.start_column, self.config.end_column + 5):
            i = get_column_letter(c)
            self.ws.column_dimensions[i].width = self.config.calendar_column_width

            for cell in self.ws[f"{c}:{c}"]:
                cell.font = self.config.main_font
                cell.alignment = self.config.center_align
        for r in range(self.config.end_row + 5):
            self.ws.row_dimensions[r].height = self.config.row_height

    def designBorder(self):
        topEdge = Border(top=thickOuterBorder)

        tlCorner = Border(top=thickOuterBorder, left=thickOuterBorder)
        trCorner = Border(top=thickOuterBorder, right=thickOuterBorder)

        dayHeader = Border(vertical=headerGrayBorder, top=headerDashBorder, bottom=headerDashBorder)
        dayHeaderL = Border(vertical=headerGrayBorder, top=headerDashBorder, bottom=headerDashBorder,
                            left=thickOuterBorder)
        dayHeaderR = Border(vertical=headerGrayBorder, top=headerDashBorder, bottom=headerDashBorder,
                            right=thickOuterBorder)

        calBodyTLC = Border(left=thickOuterBorder, top=headerDashBorder, right=calDashBorder, bottom=calDashBorder)
        calBodyTRC = Border(right=thickOuterBorder, top=headerDashBorder, left=calDashBorder, bottom=calDashBorder)
        calBodyBLC = Border(left=thickOuterBorder, top=calDashBorder, right=calDashBorder, bottom=thickOuterBorder)
        calBodyBRC = Border(right=thickOuterBorder, top=calDashBorder, left=calDashBorder, bottom=thickOuterBorder)

        calBodyL = Border(left=thickOuterBorder, top=calDashBorder, right=calDashBorder, bottom=calDashBorder)
        calBodyR = Border(right=thickOuterBorder, top=calDashBorder, left=calDashBorder, bottom=calDashBorder)

        calBodyTop = Border(top=headerDashBorder, left=calDashBorder, right=calDashBorder, bottom=calDashBorder)
        calBodyBottom = Border(top=calDashBorder, left=calDashBorder, right=calDashBorder, bottom=thickOuterBorder)

        calBody = Border(top=calDashBorder, left=calDashBorder, right=calDashBorder, bottom=calDashBorder)




        header = borderChunk(self.config.start_column, self.config.start_row, self.calendar.width, 1)
        day_head = borderChunk(self.config.start_column, self.config.day_row, self.calendar.width, 1)
        body = borderChunk(self.config.start_column, self.config.body_row, self.calendar.width, self.calendar.height)







        for row in range(start_row, start_row + 7):
            for col in range(start_column, start_column + 7):
                cell = ws.cell(column=col, row=row)
                if row == start_row:
                    cell.border = topEdge
                    if col == start_column:
                        cell.border = tlCorner
                    if col == start_column + 6:
                        cell.border = trCorner
                if row == start_row + 1:
                    cell.border = dayHeader
                    if col == start_column:
                        cell.border = dayHeaderL
                    if col == start_column + 6:
                        cell.border = dayHeaderR

                if row in range(start_row + 2, start_row + 7):
                    cell.border = calBody

                    if col == start_column:
                        cell.border = calBodyL
                    if col == start_column + 6:
                        cell.border = calBodyR

                    if row == start_row + 2:
                        cell.border = calBodyTop
                        if col == start_column:
                            cell.border = calBodyTLC
                        if col == start_column + 6:
                            cell.border = calBodyTRC
                    if row == start_row + 6:
                        cell.border = calBodyBottom
                        if col == start_column:
                            cell.border = calBodyBLC
                        if col == start_column + 6:
                            cell.border = calBodyBRC


class CalendarWriter:
    def __init__(self, config: WriterConfig, cal: _Calendar):
        self.config = config
        self.wb = self.config.wb
        self.ws = self.wb.active
        self.calendar = cal

    def write(self):
        self.ws.cell(column=self.config.start_column, row=self.config.start_row, value=self.calendar.header)

        j = 0
        for i in range(self.config.start_column, self.config.start_column+7):
            self.ws.cell(column=i, row=self.config.day_row, value=self.calendar.DAYS_COLUMNS[j])
            j+=1

        j = 0
        i = 0
        for row in range(self.config.body_row, self.config.body_row + 7):
            for col in range(self.config.start_column, self.config.start_column + 7):
                self.ws.cell(column=col, row=row, value=self.calendar.df.iloc[i, j])
            j = 0
            i+= 1

    def style_page(self):
        self.ws.page_setup.orientation = self.config.page_orientation
        self.ws.page_margins = self.config.page_margins
        self.ws.page_setup.paperSize = self.config.paper_Size

    def save(self):
        self.wb.save( self.config.name)


class ExcelBookGenerator:
    def __init__(self, cal: _Calendar, writer: CalendarWriter, styler:CalendarStyler, config: CalendarConfig):
        self.calendar = cal
        self.writer = writer
        self.styler = styler
        self.config = config

    def generate(self):
        self.writer.write()






thickOuterBorder= Side(border_style='thick', color=black)
headerDashBorder = Side(border_style='mediumDashed', color=black)
headerGrayBorder = Side(border_style='thin', color=lightGray)
calDashBorder = Side(border_style='dashed', color=black)


leftEdge = Border(left=thickOuterBorder)
rightEdge = Border(right=thickOuterBorder)
bottomEdge = Border(bottom=thickOuterBorder)
topEdge = Border(top=thickOuterBorder)
dayHeader = Border(vertical= headerGrayBorder, top = headerDashBorder, bottom=headerDashBorder)
calBody = Border(vertical= calDashBorder, horizontal = calDashBorder)

border = Border(left=Side(border_style=None,
                          color='FF000000'),
                right=Side(border_style=None,
                           color='FF000000'),
                top=Side(border_style=None,
                         color='FF000000'),
                bottom=Side(border_style=None,
                            color='FF000000'),
                outline=Side(border_style=None,
                             color='FF000000'),
                vertical=Side(border_style=None,
                              color='FF000000'),
                horizontal=Side(border_style=None,
                               color='FF000000')
                )
number_format = 'General'




def getRuMonth(_year, _month):
    cal = calendar.LocaleTextCalendar(locale='ru')
    return cal.formatmonthname(_year, _month, 0, withyear=False)


def localeCal(_year, _month):
    cal = calendar.LocaleTextCalendar(locale='ru')
    month = cal.formatmonth(_year, _month)
    return cal, month

def dateCal(_year, _month):

def calToList(month):
    monthList = []
    for w in month:
        week = []
        for d in w:
            week.append(d.day)
        monthList.append(week)
    return monthList

def CalToDf(month):


def isNowMonth(month=0):
    if month == _month:
        return True
    else:
        return False









def designCal(ws, oddIndex):

def widthHeightFormat(sheet):
    sheet.column_dimensions['B'].width = 60
    sheet.row_dimensions[1].height = 4


def main():




main()

print(getRuMonth(_year, _month))
