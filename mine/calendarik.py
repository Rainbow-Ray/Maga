import calendar
import openpyxl as ex
import pandas as pd
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image


start_column = 2
start_row = 2

black = 'FF000000'
lightGray = 'd8d8d8'
gray = 'a5a5a5'

mainFont = Font(name='Century Gothic',
                size=20,
                bold=False,
                italic=False,
                vertAlign='baseline',
                underline='none',
                strike=False,
                color='FF000000')

oddFont = Font(name='Century Gothic',
                size=20,
                bold=False,
                italic=False,
                vertAlign='baseline',
                underline='none',
                strike=False,
                color=gray)


thickOuterBorder= Side(border_style='thick', color=black)
headerDashBorder = Side(border_style='mediumDashed', color=black)
headerGrayBorder = Side(border_style='thin', color=lightGray)
calDashBorder = Side(border_style='dashed', color=black)


leftEdge = Border(left=thickOuterBorder)
rightEdge = Border(right=thickOuterBorder)
bottomEdge = Border(bottom=thickOuterBorder)
topEdge = Border(top=thickOuterBorder)
dayHeader = Border(vertical= headerGrayBorder, top = headerDashBorder, bottom=headerDashBorder)
calBody = Border(vertical= calDashBorder, horizontal = calDashBorder)

border = Border(left=Side(border_style=None,
                          color='FF000000'),
                right=Side(border_style=None,
                           color='FF000000'),
                top=Side(border_style=None,
                         color='FF000000'),
                bottom=Side(border_style=None,
                            color='FF000000'),
                outline=Side(border_style=None,
                             color='FF000000'),
                vertical=Side(border_style=None,
                              color='FF000000'),
                horizontal=Side(border_style=None,
                               color='FF000000')
                )
centerAlign=Alignment(horizontal='center',
                    vertical='center',
                    text_rotation=0,
                    wrap_text=False,
                    shrink_to_fit=False,
                    indent=0)
number_format = 'General'



_year = 2025
_month = 11
daysColumn = {
    0: 'Пн',
    1: 'Вт',
    2: 'Ср',
    3: 'Чт',
    4: 'Пт',
    5: 'Сб',
    6: 'Вс',
}



def getRuMonth(_year, _month):
    cal = calendar.LocaleTextCalendar(locale='ru')
    return cal.formatmonthname(_year, _month, 0, withyear=False)


def localeCal(_year, _month):
    cal = calendar.LocaleTextCalendar(locale='ru')
    month = cal.formatmonth(_year, _month)
    return cal, month

def dateCal(_year, _month):
    cal = calendar.Calendar()
    month = cal.monthdatescalendar(_year, _month)
    return cal, month

def calToList(month):
    monthList = []
    for w in month:
        week = []
        for d in w:
            week.append(d.day)
        monthList.append(week)
    return monthList

def CalToDf(month):
    df = pd.DataFrame(month)
    df = df.rename(columns= daysColumn)
    return df


def isNowMonth(month=0):
    if month == _month:
        return True
    else:
        return False

def oddMonthIndex(df, row):
    s = -1
    f = -1
    for i in range(7):
        c = df.iloc[row,i]
        if not isNowMonth(c.month):
            if s == -1:
                s = i
            f = i

    return s,f

def getOddMonth(df):
    monthNow = _month
    oldS, oldF = oddMonthIndex(df, 0)
    newS, newF = oddMonthIndex(df, -1)

    return {'Old':[oldS, oldF], 'New':(newS, newF)}


def calFormatToDay(df):
    for h in df.columns:
        df[h] = df.apply(lambda cell: cell[h].day, axis=1)
    return df

def writeCalToSheet(ws, header, column, df):
    ws.cell(column=start_column, row=start_row, value=header)
    j = 0
    for i in range(2,9):
        ws.cell(column=i, row=start_row+1, value=column[i-2])
        j+=1

    for row in range(start_row+2,start_row+7):
        for col in range(start_column,start_column+7):
            ws.cell(column=col, row=row, value= df.iloc[row-start_row-2, col-start_column])


def fontOddMonth(ws, start, finish, style):
    for i in range(start, finish):
        ws.cell(column=i, row=start_row+2).font = style

def designBorder(ws):
    leftEdge = Border(left=thickOuterBorder)
    rightEdge = Border(right=thickOuterBorder)
    bottomEdge = Border(bottom=thickOuterBorder)
    topEdge = Border(top=thickOuterBorder)

    tlCorner = Border(top=thickOuterBorder, left=thickOuterBorder)
    trCorner = Border(top=thickOuterBorder, right=thickOuterBorder)
    blCorner = Border(bottom=thickOuterBorder, left=thickOuterBorder)
    brCorner = Border(bottom=thickOuterBorder, right=thickOuterBorder)

    dayHeader = Border(vertical=headerGrayBorder, top=headerDashBorder, bottom=headerDashBorder)
    dayHeaderL = Border(vertical=headerGrayBorder, top=headerDashBorder, bottom=headerDashBorder, left=thickOuterBorder)
    dayHeaderR = Border(vertical=headerGrayBorder, top=headerDashBorder, bottom=headerDashBorder, right=thickOuterBorder)

    calBodyTLC = Border(left =thickOuterBorder, top=headerDashBorder,  right=calDashBorder, bottom=calDashBorder)
    calBodyTRC = Border(right =thickOuterBorder, top=headerDashBorder,  left=calDashBorder, bottom=calDashBorder)
    calBodyBLC = Border(left =thickOuterBorder, top=calDashBorder,  right=calDashBorder, bottom=thickOuterBorder)
    calBodyBRC = Border(right =thickOuterBorder, top=calDashBorder,  left=calDashBorder, bottom=thickOuterBorder)

    calBodyL = Border(left =thickOuterBorder, top=calDashBorder,  right=calDashBorder, bottom=calDashBorder)
    calBodyR = Border(right =thickOuterBorder, top=calDashBorder,  left=calDashBorder, bottom=calDashBorder)

    calBodyTop = Border(top=headerDashBorder,left=calDashBorder, right=calDashBorder, bottom=calDashBorder)
    calBodyBottom = Border(top=calDashBorder,left=calDashBorder, right=calDashBorder, bottom=thickOuterBorder)

    calBody = Border(top=calDashBorder, left=calDashBorder, right=calDashBorder, bottom=calDashBorder)

    for row in range(start_row, start_row+7):
        for col in range(start_column, start_column+7):
            cell = ws.cell(column=col, row=row)
            if row == start_row:
                cell.border = topEdge
                if col == start_column:
                    cell.border = tlCorner
                if col == start_column+6:
                    cell.border = trCorner
            if row == start_row+1:
                cell.border = dayHeader
                if col == start_column:
                    cell.border = dayHeaderL
                if col== start_column+6:
                    cell.border = dayHeaderR

            if row in range(start_row+2,start_row+7):
                cell.border = calBody

                if col == start_column:
                    cell.border = calBodyL
                if col == start_column+6:
                    cell.border = calBodyR

                if row == start_row + 2:
                    cell.border = calBodyTop
                    if col == start_column:
                        cell.border = calBodyTLC
                    if col == start_column+6:
                        cell.border = calBodyTRC
                if row == start_row+6:
                    cell.border = calBodyBottom
                    if col == start_column:
                        cell.border = calBodyBLC
                    if col == start_column+6:
                        cell.border = calBodyBRC


def designCal(ws, oddIndex):
    s = get_column_letter(start_column)
    e = get_column_letter(start_column+6)
    ws.merge_cells(f'{s}{start_row}:{e}{start_row}')

    ws.column_dimensions['A'].width = 25.6

    for c in range(2, 15):
        i = get_column_letter(c)
        ws.column_dimensions[i].width = 13.6

        for cell in ws[f"{c}:{c}"]:
            cell.font = mainFont
            cell.alignment = centerAlign

    for r in range(15):
        ws.row_dimensions[r].height = 59.40

    if oddIndex['Old'][0] != -1:
        fontOddMonth(ws, oddIndex['Old'][0] + 2, oddIndex['Old'][1] + 3, oddFont)
    if oddIndex['New'][0] != -1:
        fontOddMonth(ws, oddIndex['New'][0] + 2, oddIndex['New'][1] + 3, oddFont)

    designBorder(ws)

def widthHeightFormat(sheet):
    sheet.column_dimensions['B'].width = 60
    sheet.row_dimensions[1].height = 4


def main():
    cal, month = dateCal(_year, _month)
    df = CalToDf(month)

    oddIndex = getOddMonth(df)
    wb = ex.Workbook()
    ws = wb.active
    header = getRuMonth(_year, _month)
    df = calFormatToDay(df)
    column = (df.columns.values).tolist()

    writeCalToSheet(ws, header, column, df)
    designCal(ws,oddIndex)

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_margins = PageMargins(left=0, right=0, top=0, bottom=0, header=0, footer=0)
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    img = Image('4.jpg')
    ws.add_image(img, 'A1')

    wb.save("pandas_openpyxl.xlsx")





main()

print(getRuMonth(_year, _month))


# def tablde():
# start_date = 10
# end_date = 10

#     tb = (
#         GT(df)
#         .tab_header(title="Октябрь")
#         .fmt_integer(columns=['Пн','Вт','Ср', 'Чт', 'Пт','Сб','Вс'], compact=True)
#         .tab_style(locations=loc.header(),
#                    style=[style.text(font='overdoze sans'),
#                           style.borders(sides='all', color='#000000', style='dashed', weight='2px')
#                           ])
#
#         .tab_style(locations=loc.body(columns=['Пн', 'Вт'], rows=[0]),
#                    style = [
#                             style.text(color="gray", weight="thin", )])
#         .tab_style(locations=loc.body(columns=['Сб','Вс'], rows=[4]),
#                    style = [style.text(color="gray", weight="thin")])
#         .tab_style(locations=loc.body(),
#                    style = [style.text( align='center', v_align='middle'), ])
#     )
#     tb.show()
